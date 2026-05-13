import pandas as pd
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_reconciliation_excel(report_data: Dict[str, Any]) -> io.BytesIO:
    """
    Generates a high-fidelity, comprehensive Excel reconciliation report.
    Includes Overview, Matched, Possible, Unmatched Bank, Unmatched Xero, and Audit Trail.
    """
    output = io.BytesIO()
    
    # --- Data Extraction ---
    summary = report_data.get("summary", {})
    buckets = report_data.get("buckets", {})
    
    matched_items = buckets.get("matched", [])
    possible_items = buckets.get("possible", [])
    unmatched_bank_items = buckets.get("unmatched_bank", [])
    unmatched_xero_items = buckets.get("unmatched_xero", [])

    # --- 1. Summary Sheet (KPIs) ---
    summary_data = [
        {"METRIC": "UPLOAD OVERVIEW", "VALUE": "", "DETAILS": ""},
        {"METRIC": "Total Bank Transactions", "VALUE": summary.get("total_bank_rows", 0), "DETAILS": f"${summary.get('matched_amount', 0) + summary.get('unmatched_bank_amount', 0) + summary.get('possible_amount', 0):,.2f}"},
        {"METRIC": "Total Xero Invoices", "VALUE": summary.get("total_xero_invoices", 0), "DETAILS": ""},
        {"METRIC": "", "VALUE": "", "DETAILS": ""},
        {"METRIC": "RECONCILIATION BREAKDOWN", "VALUE": "COUNT", "DETAILS": "TOTAL AMOUNT ($)"},
        {"METRIC": "Successfully Matched", "VALUE": summary.get("matched_count", 0), "DETAILS": f"${summary.get('matched_amount', 0):,.2f}"},
        {"METRIC": "Possible Matches (Pending)", "VALUE": summary.get("possible_count", 0), "DETAILS": f"${summary.get('possible_amount', 0):,.2f}"},
        {"METRIC": "Outstanding Bank (Unmatched)", "VALUE": summary.get("unmatched_bank_count", 0), "DETAILS": f"${summary.get('unmatched_bank_amount', 0):,.2f}"},
        {"METRIC": "Outstanding Xero (Unmatched)", "VALUE": summary.get("unmatched_xero_count", 0), "DETAILS": f"${summary.get('unmatched_xero_amount', 0):,.2f}"},
    ]
    df_summary = pd.DataFrame(summary_data)

    # --- 2. Matched Details Sheet ---
    matched_list = []
    for match in matched_items:
        bank_transaction = match.get("bank_transaction", {})
        xero_invoice = match.get("xero_invoice", {})
        matched_list.append({
            "DATE": bank_transaction.get("transaction_date"),
            "BANK DESCRIPTION": bank_transaction.get("description"),
            "BANK AMOUNT": bank_transaction.get("amount"),
            "LINKED INVOICE": xero_invoice.get("InvoiceNumber"),
            "CONTACT": xero_invoice.get("Contact", {}).get("Name"),
            "CONFIDENCE": f"{match.get('confidence')}%",
            "TYPE": "Manual" if match.get("is_manual") else "Auto"
        })
    df_matched = pd.DataFrame(matched_list) if matched_list else pd.DataFrame([{"INFO": "No matched items recorded."}])

    # --- 3. Possible Matches Sheet ---
    possible_list = []
    for possible_match in possible_items:
        bank_transaction = possible_match.get("bank_transaction", {})
        xero_invoice = possible_match.get("xero_invoice", {})
        possible_list.append({
            "DATE": bank_transaction.get("transaction_date"),
            "BANK DESCRIPTION": bank_transaction.get("description"),
            "BANK AMOUNT": bank_transaction.get("amount"),
            "SUGGESTED INVOICE": xero_invoice.get("InvoiceNumber"),
            "SUGGESTED CONTACT": xero_invoice.get("Contact", {}).get("Name"),
            "MATCH SCORE": f"{possible_match.get('confidence')}%",
            "REASON": "Ambiguous" if possible_match.get("is_ambiguous") else "Low Confidence"
        })
    df_possible = pd.DataFrame(possible_list) if possible_list else pd.DataFrame([{"INFO": "No possible matches found."}])

    # --- 4. Unmatched Bank Sheet ---
    unmatched_bank_list = []
    for bank_item in unmatched_bank_items:
        unmatched_bank_list.append({
            "DATE": bank_item.get("transaction_date"),
            "DESCRIPTION": bank_item.get("description"),
            "AMOUNT": bank_item.get("amount")
        })
    df_bank = pd.DataFrame(unmatched_bank_list) if unmatched_bank_list else pd.DataFrame([{"INFO": "No unmatched bank items found."}])

    # --- 5. Unmatched Xero Sheet ---
    unmatched_xero_list = []
    for xero_item in unmatched_xero_items:
        unmatched_xero_list.append({
            "INVOICE #": xero_item.get("InvoiceNumber"),
            "CONTACT": xero_item.get("Contact", {}).get("Name"),
            "DATE": xero_item.get("DateString") or xero_item.get("Date"),
            "AMOUNT": xero_item.get("Total"),
            "STATUS": xero_item.get("Status")
        })
    df_xero = pd.DataFrame(unmatched_xero_list) if unmatched_xero_list else pd.DataFrame([{"INFO": "No unmatched Xero invoices found."}])

    # --- 6. Audit Trail Sheet ---
    audit_list = []
    for match in matched_items:
        if match.get("is_manual"):
            bank_transaction = match.get("bank_transaction", {})
            xero_invoice = match.get("xero_invoice", {})
            audit_list.append({
                "USER CONFIRMATION TIMESTAMP": bank_transaction.get("reconciled_at") or "Existing Record",
                "BANK TRANSACTION": bank_transaction.get("description"),
                "LINKED XERO INVOICE": f"{xero_invoice.get('InvoiceNumber')} ({xero_invoice.get('Contact', {}).get('Name')})",
                "AMOUNT": bank_transaction.get("amount"),
                "RESULT": "Manual Link Approved"
            })
    df_audit = pd.DataFrame(audit_list) if audit_list else pd.DataFrame([{"INFO": "No manual approvals recorded."}])

    # --- Write and Style ---
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name="Overview", index=False)
        df_matched.to_excel(writer, sheet_name="Matched Items", index=False)
        df_possible.to_excel(writer, sheet_name="Possible Matches", index=False)
        df_bank.to_excel(writer, sheet_name="Unmatched Bank", index=False)
        df_xero.to_excel(writer, sheet_name="Unmatched Xero", index=False)
        df_audit.to_excel(writer, sheet_name="Audit Trail", index=False)

        # Styling
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style='thin', color="D1D5DB"), 
                        right=Side(style='thin', color="D1D5DB"), 
                        top=Side(style='thin', color="D1D5DB"), 
                        bottom=Side(style='thin', color="D1D5DB"))

        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    cell.border = border
                    if cell.row > 1 and cell.row % 2 == 0:
                        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws.column_dimensions[column].width = min(max_length + 3, 60)

    output.seek(0)
    return output
